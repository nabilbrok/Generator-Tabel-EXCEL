import subprocess
import sys

# Fungsi untuk menginstal pustaka jika belum ada
def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# Daftar pustaka yang dibutuhkan
required_packages = ['pandas', 'openpyxl']

# Memeriksa dan menginstal setiap pustaka yang dibutuhkan
for package in required_packages:
    try:
        __import__(package)
    except ImportError:
        print(f"{package} tidak ditemukan, menginstal...")
        install(package)

import random
import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import Border, Side, Alignment, Font

# Daftar perangkat elektronik umum dengan rentang watt yang lebih realistis
daftar_perangkat = [
    {"nama": "Lampu LED", "watt_min": 5, "watt_max": 15, "harus_digunakan": True},
    {"nama": "Lampu Neon", "watt_min": 20, "watt_max": 40, "harus_digunakan": True},
    {"nama": "TV LED", "watt_min": 60, "watt_max": 150, "harus_digunakan": False},
    {"nama": "Kulkas", "watt_min": 100, "watt_max": 250, "harus_digunakan": True},
    {"nama": "AC 1/2 PK", "watt_min": 400, "watt_max": 500, "harus_digunakan": False},
    {"nama": "Kipas Angin", "watt_min": 30, "watt_max": 70, "harus_digunakan": True},
    {"nama": "Komputer Desktop", "watt_min": 200, "watt_max": 600, "harus_digunakan": False},
    {"nama": "Lampu Meja", "watt_min": 10, "watt_max": 20, "harus_digunakan": True},
    {"nama": "Telepon Nirkabel", "watt_min": 1, "watt_max": 5, "harus_digunakan": True},
    {"nama": "Charger Ponsel", "watt_min": 5, "watt_max": 20, "harus_digunakan": True},
    {"nama": "Modem Wi-Fi", "watt_min": 10, "watt_max": 20, "harus_digunakan": True},
    {"nama": "Kipas Angin Dinding", "watt_min": 50, "watt_max": 100, "harus_digunakan": False},
    {"nama": "Dispenser", "watt_min": 150, "watt_max": 300, "harus_digunakan": True},
    {"nama": "Kipas Angin Dinding", "watt_min": 50, "watt_max": 100, "harus_digunakan": False},
    {"nama": "Dispenser", "watt_min": 150, "watt_max": 300, "harus_digunakan": True},
    {"nama": "Setrika", "watt_min": 1000, "watt_max": 2000, "harus_digunakan": False},
    {"nama": "Mesin Cuci", "watt_min": 300, "watt_max": 800, "harus_digunakan": False},
    {"nama": "Pengering Rambut", "watt_min": 800, "watt_max": 1800, "harus_digunakan": False},
    {"nama": "Pemanggang Roti", "watt_min": 800, "watt_max": 1500, "harus_digunakan": False},
    {"nama": "Oven Listrik", "watt_min": 1000, "watt_max": 5000, "harus_digunakan": False},
    {"nama": "Rice Cooker", "watt_min": 300, "watt_max": 1000, "harus_digunakan": True},
    {"nama": "Penyedot Debu", "watt_min": 500, "watt_max": 1500, "harus_digunakan": False},
    {"nama": "Sound System", "watt_min": 50, "watt_max": 200, "harus_digunakan": False},
    {"nama": "Proyektor", "watt_min": 200, "watt_max": 400, "harus_digunakan": False},
    {"nama": "Mesin Kopi", "watt_min": 600, "watt_max": 1200, "harus_digunakan": False},
    {"nama": "Jam Digital", "watt_min": 1, "watt_max": 5, "harus_digunakan": True},
]

# Fungsi untuk memilih perangkat yang digunakan
# Fungsi untuk memilih perangkat yang digunakan
def pilih_perangkat(jumlah):
    # Memastikan perangkat yang harus digunakan selalu ada dalam daftar
    perangkat_harus_digunakan = [p for p in daftar_perangkat if p['harus_digunakan']]
    total_harus_digunakan = len(perangkat_harus_digunakan)

    # Memastikan jumlah yang diminta cukup untuk perangkat yang harus digunakan
    if jumlah < total_harus_digunakan:
        raise ValueError(f"Jumlah perangkat harus minimal {total_harus_digunakan} (termasuk perangkat yang harus digunakan).")

    perangkat_acak = [p for p in daftar_perangkat if not p['harus_digunakan']]
    jumlah_perangkat_acak = jumlah - total_harus_digunakan

    # Mengambil perangkat acak sesuai jumlah yang diminta
    perangkat_terpilih = random.sample(perangkat_acak, jumlah_perangkat_acak) if jumlah_perangkat_acak > 0 else []

    return perangkat_harus_digunakan + perangkat_terpilih


# Fungsi untuk menghitung kWh
def hitung_kwh(watt, jam_per_hari):
    return (watt * jam_per_hari * 30) / 1000  # Menghitung kWh per bulan

# Fungsi untuk menyesuaikan jam agar total kWh sesuai
def sesuaikan_jam(perangkat, total_kwh_bulan):
    total_kwh_saat_ini = sum(hitung_kwh(p['watt'], p['jam_per_hari']) for p in perangkat)
    faktor_penyesuaian = total_kwh_bulan / total_kwh_saat_ini if total_kwh_saat_ini > 0 else 1

    for p in perangkat:
        p['jam_per_hari'] = round(p['jam_per_hari'] * faktor_penyesuaian)  # Bulatkan jam per hari

    return perangkat

# Fungsi untuk membuat file Excel
def buat_excel(nama_file, data_perangkat):
    df = pd.DataFrame(data_perangkat)

    # Hitung konsumsi listrik (kWh) dan tambahkan ke DataFrame
    df['Konsumsi Listrik (kWh)'] = df.apply(lambda row: hitung_kwh(row['Daya (Watt)'], row['Lama Penggunaan (jam/hari)']), axis=1)

    # Hanya simpan perangkat yang memiliki penggunaan listrik
    df = df[df['Konsumsi Listrik (kWh)'] > 0]

    # Tambahkan kolom total kWh
    total_kwh = df['Konsumsi Listrik (kWh)'].sum()
    total_row = pd.DataFrame({"Nama Perangkat": ["Total"], "Daya (Watt)": ["-"], "Lama Penggunaan (jam/hari)": ["-"], "Konsumsi Listrik (kWh)": [total_kwh]})
    df = pd.concat([df, total_row], ignore_index=True)

    # Membuat tabel di Excel
    with pd.ExcelWriter(nama_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Penggunaan Listrik')
        sheet = writer.sheets['Penggunaan Listrik']
        buat_tabel_terstruktur(sheet)

# Fungsi untuk melakukan auto-resize kolom berdasarkan isi
def auto_resize_columns(sheet):
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Mengambil huruf kolom
        for cell in col:
            try:
                # Menghitung panjang maksimum di setiap kolom
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Menetapkan lebar kolom berdasarkan panjang maksimum
        adjusted_width = max_length + 2  # Menambahkan sedikit padding
        sheet.column_dimensions[col_letter].width = adjusted_width

# Fungsi untuk menambahkan style pada tabel di Excel
def buat_tabel_terstruktur(sheet):
    # Definisi untuk border
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Menerapkan border, alignment, dan bold pada header
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)

    # Menambahkan border untuk semua data
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Panggil fungsi auto-resize untuk mengatur lebar kolom otomatis
    auto_resize_columns(sheet)

# Fungsi utama
def main():
    try:
        # Menentukan jumlah perangkat yang harus digunakan
        jumlah_perangkat_harus_digunakan = sum(1 for p in daftar_perangkat if p['harus_digunakan'])
        
        jumlah_perangkat = int(input("Masukkan jumlah perangkat di rumah (minimal {}): ".format(jumlah_perangkat_harus_digunakan)))
        if jumlah_perangkat <= 0:
            print("Jumlah perangkat harus lebih dari 0.")
            return

        perangkat_terpilih = pilih_perangkat(jumlah_perangkat)

        for perangkat in perangkat_terpilih:
            perangkat['watt'] = random.randint(perangkat['watt_min'], perangkat['watt_max'])
            perangkat['jam_per_hari'] = round(random.uniform(0.1, 24), 0)  # Bulatkan jam ke bilangan bulat

        total_kwh_bulan = float(input("Masukkan total kWh yang digunakan dalam sebulan: "))
        perangkat_terpilih = sesuaikan_jam(perangkat_terpilih, total_kwh_bulan)

        # Buat list perangkat yang akan dimasukkan ke DataFrame
        data_perangkat = [
            {"Nama Perangkat": p['nama'], "Daya (Watt)": p['watt'], "Lama Penggunaan (jam/hari)": p['jam_per_hari']}
            for p in perangkat_terpilih if p['jam_per_hari'] > 0  # Hanya memasukkan perangkat yang digunakan
        ]

        # Buat folder jika belum ada
        folder = "file_excel"
        if not os.path.exists(folder):
            os.makedirs(folder)

        # Menambahkan timestamp agar file tidak overwrite
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nama_file = f"{folder}/Penggunaan_Listrik_{timestamp}.xlsx"

        # Buat Excel
        buat_excel(nama_file, data_perangkat)

        print(f"File '{nama_file}' berhasil dibuat.")
    except ValueError as ve:
        print(f"Input tidak valid: {ve}")

if __name__ == "__main__":
    main()